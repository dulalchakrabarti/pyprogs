from PIL import Image
import openpyxl

im = Image.open('meme.jpg', 'r')

### Adjustible parameters ###
max_width = 350
sheet_zoom = 15             # Zoom level of excel sheet
column_size = 2
row_size = 10
img_left_margin = 10        
img_top_margin = 10         
img_right_margin = 10       
img_bottom_margin = 10      
bg_col = 'FFFFFF'           # Background colour - default is white
black_and_white = 0         # Set to 1 if you want black and white image
#############################

width = max_width
height = int(width*(im.size[1]/im.size[0]))
im = im.resize((width, height))
px = im.load()

wb = openpyxl.Workbook()
ws = wb.active

# Setting cell seizes
for i in range(0, height+img_top_margin+img_bottom_margin):
    t = ws.cell(row = i+1, column = 1)
    ws.row_dimensions[t.row].height = row_size
for j in range(0, width+img_left_margin+img_right_margin):
    t = ws.cell(row = 1, column = j+1)
    ws.column_dimensions[t.coordinate.replace(str(t.row),'')].width = column_size

### Optional - setting a background color ###    
for i in range(0, width+img_left_margin+img_right_margin):
    for j in range(0, height+img_top_margin+img_bottom_margin):
        a = ws.cell(row = j+1, column = i+1)
        a.fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = bg_col, end_color = bg_col)
#############################################
        
for i in range(0, width):
    for j in range(0, height):
        a = ws.cell(row = j+img_top_margin, column = i+img_left_margin)
        col = px[i, j]
        c1 = col[0]
        c2 = col[1]
        c3 = col[2]
        
        if black_and_white == 1:
            bw = int(0.2126*c1 + 0.7152*c2 + 0.0722*c3)
            val = '{:02X}'.format(bw)
            c = val + val + val
        else:
            c = '{:02X}{:02X}{:02X}'.format(c1, c2, c3)

        c_fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = c, end_color = c)
        a.fill = c_fill

ws.sheet_view.zoomScale = sheet_zoom
wb.save(filename = 'meme_pic.xlsx')
